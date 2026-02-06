import io
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook

from utils import (
    find_headers,
    extract_stations,
    extract_train_columns,
    parse_time,
    format_time_decimal,
    apply_midnight_correction,
)


def read_workbook(file_bytes: bytes) -> Tuple[List[str], Dict[str, pd.DataFrame]]:
    """Read all sheets with merged cells expanded so every cell in a merged range
    carries the top-left value. Returns {name: DataFrame}.

    The resulting DataFrames keep Python None for empty cells (so pd.isna works),
    and preserve original types where possible.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    sheet_names = wb.sheetnames
    sheets: Dict[str, pd.DataFrame] = {}

    for name in sheet_names:
        ws = wb[name]

        # Expand merged cells: copy top-left value into all cells of the merged range
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            min_row, max_row = mr.min_row, mr.max_row
            min_col, max_col = mr.min_col, mr.max_col
            top_left_value = ws.cell(row=min_row, column=min_col).value
            # Unmerge first to avoid 'MergedCell' read-only value error
            try:
                ws.unmerge_cells(range_string=str(mr))
            except Exception:
                pass
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_left_value

        # Determine effective used area (trim trailing empty rows/cols)
        effective_max_row = 0
        effective_max_col = 0
        row_values: List[List[Optional[object]]] = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = list(row) if row is not None else []
            last_nonempty_col = 0
            for c_idx, val in enumerate(values, start=1):
                if val is not None and (not isinstance(val, str) or val != ""):
                    last_nonempty_col = c_idx
            if last_nonempty_col > 0:
                effective_max_row = r_idx
                if last_nonempty_col > effective_max_col:
                    effective_max_col = last_nonempty_col
            row_values.append(values)

        # Build rectangular data limited to effective bounds
        trimmed_rows: List[List[Optional[object]]] = []
        for values in row_values[:effective_max_row]:
            row_list = list(values[:effective_max_col])
            if len(row_list) < effective_max_col:
                row_list.extend([None] * (effective_max_col - len(row_list)))
            trimmed_rows.append(row_list)

        df = pd.DataFrame(trimmed_rows)
        sheets[name] = df

    return sheet_names, sheets


def extract_excel_data(sheet_names: List[str], sheets: Dict[str, pd.DataFrame]):
    """Extract station map from first sheet and per-sheet trains data.

    Returns a dict with keys:
    - station_map: Dict[str, float]  # station -> km from the first sheet
    - station_check: Dict[str, any]  # {'ok': bool, 'mismatches': List[str]}
    - sheets_data: List[Dict[str, any]]  # [{'sheet': name, 'trains': List[...]}]
    """
    if not sheet_names:
        return {
            "station_map": {},
            "station_check": {"ok": False, "mismatches": ["Workbook has no sheets."]},
            "sheets_data": [],
        }

    first_sheet = sheet_names[0]
    df_first = sheets[first_sheet]
    pos_first = find_headers(df_first)

    if not all(
        pos_first.get(k) is not None
        for k in ("station_start_row", "station_end_row", "station_col", "km_col")
    ):
        raise ValueError("Missing station headers in the first sheet.")

    stations_first = extract_stations(
        df_first,
        start_row=pos_first["station_start_row"],
        end_row=pos_first["station_end_row"],
        station_col=pos_first["station_col"],
        km_col=pos_first["km_col"],
    )

    # station_map based on the first sheet
    station_to_km = {station: km for km, station, _ in stations_first}
    reference_station_set = set(station_to_km.keys())

    mismatches: List[str] = []
    sheets_data: List[Dict[str, any]] = []
    station_maps: Dict[str, Dict[str, float]] = {sheet_names[0]: station_to_km}

    for sheet in sheet_names:
        df = sheets[sheet]
        pos = find_headers(df)

        # Extract and verify station list for the sheet (order-independent)
        if all(
            pos.get(k) is not None
            for k in ("station_start_row", "station_end_row", "station_col", "km_col")
        ):
            stations = extract_stations(
                df,
                start_row=pos["station_start_row"],
                end_row=pos["station_end_row"],
                station_col=pos["station_col"],
                km_col=pos["km_col"],
            )
            # per-sheet station map
            station_maps[sheet] = {s: km for km, s, _ in stations}
            sheet_station_set = {s for _, s, _ in stations}
            if sheet_station_set != reference_station_set:
                only_in_sheet = sorted(sheet_station_set - reference_station_set)
                only_in_ref = sorted(reference_station_set - sheet_station_set)
                if only_in_sheet:
                    mismatches.append(
                        f"Sheet '{sheet}' has stations not in reference: {only_in_sheet}"
                    )
                if only_in_ref:
                    mismatches.append(
                        f"Sheet '{sheet}' misses stations from reference: {only_in_ref}"
                    )
        else:
            mismatches.append(f"Sheet '{sheet}' is missing station headers.")
            stations = []
            station_maps[sheet] = {}

        # Extract trains
        trains_list: List[Dict[str, any]] = []
        if pos.get("train_row") is not None and stations:
            train_columns = extract_train_columns(
                df,
                pos["train_row"],
                station_start_row=pos.get("station_start_row"),
                station_end_row=pos.get("station_end_row"),
            )

            # Build entries: train_number - station - km (from this sheet) - time (HH:MM or HH:MM (+d))
            # Collect raw times per train first, then apply midnight correction
            for train_nr, col in train_columns.items():
                raw_entries = []  # (station_name, km_ref, raw_time)
                for km_ref, station_name, row_idx in stations:
                    raw_val = df.iat[row_idx, col]
                    t = parse_time(raw_val)
                    if t is None:
                        continue
                    raw_entries.append((station_name, float(km_ref), float(t)))

                if not raw_entries:
                    continue

                # Apply midnight correction to the sequence of raw times
                raw_times = [e[2] for e in raw_entries]
                corrected_times = apply_midnight_correction(raw_times)

                for i, (station_name, km_ref, _raw_t) in enumerate(raw_entries):
                    corrected_t = corrected_times[i]
                    trains_list.append(
                        {
                            "train_number": str(train_nr),
                            "station": station_name,
                            "km": km_ref,
                            "time": format_time_decimal(corrected_t),
                            "time_decimal": corrected_t,
                        }
                    )
        else:
            # No trains found or no stations available to map
            pass

        sheets_data.append({"sheet": sheet, "trains": trains_list})

    station_check = {"ok": len(mismatches) == 0, "mismatches": mismatches}

    return {
        "station_map": station_to_km,
        "station_maps": station_maps,
        "station_check": station_check,
        "sheets_data": sheets_data,
    }


def read_and_store_in_session(file_bytes: bytes, session_state) -> None:
    """High-level helper to read workbook, extract data, and store in session_state."""
    sheet_names, sheets = read_workbook(file_bytes)
    data = extract_excel_data(sheet_names, sheets)
    session_state["station_map"] = data["station_map"]
    session_state["station_maps"] = data.get("station_maps", {})
    session_state["station_check"] = data["station_check"]
    session_state["sheets_data"] = data["sheets_data"]


