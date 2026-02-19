"""Export logic extracted from app.py: XLSX timetable, vehicle circuits, project JSON."""
from __future__ import annotations

import datetime as dt
import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from utils import parse_time, format_time_hhmm, normalize


def _hhmm_from_any(val: Any) -> str:
    try:
        if isinstance(val, (int, float)):
            d = float(val)
        else:
            d = parse_time(val)
        if d is None:
            return ""
        return format_time_hhmm(d)
    except Exception:
        return ""


def _decimal_to_time(d: float) -> dt.time:
    h = int(d) % 24
    m = int(round((d % 1) * 60))
    if m == 60:
        h = (h + 1) % 24
        m = 0
    return dt.time(h, m)


def build_excel_bytes(session: Any) -> bytes:
    sheets_data = session.get("sheets_data", [])
    station_maps_all = session.get("station_maps", {})

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    for entry in sheets_data:
        sheet_name = str(entry.get("sheet"))
        ws = wb.create_sheet(title=sheet_name[:31] or "Arkusz")

        ws["E3"] = "numer pociągu"
        ws["D11"] = "km"
        ws["E11"] = "ze stacji"
        ws["F11"] = "p/o"

        station_map_sheet = station_maps_all.get(sheet_name, {})
        stations_sorted = sorted(station_map_sheet.items(), key=lambda kv: kv[1])

        start_row = 12

        trains_list = entry.get("trains", [])
        train_nums = list(dict.fromkeys(str(t.get("train_number")) for t in trains_list))
        for j, tn in enumerate(train_nums):
            ws.cell(row=3, column=7 + j, value=tn)

        key_to_time: dict[tuple, dict] = {}
        key_to_time_norm: dict[tuple, dict] = {}
        for rec in trains_list:
            st_name = rec.get("station")
            tn = str(rec.get("train_number"))
            t_fmt = _hhmm_from_any(rec.get("time_decimal") if rec.get("time_decimal") is not None else rec.get("time"))
            rec_stop_type = rec.get("stop_type", "p")
            key_to_time.setdefault((st_name, tn), {})[rec_stop_type] = t_fmt
            key_to_time_norm.setdefault((normalize(str(st_name)), tn), {})[rec_stop_type] = t_fmt

        def _get_time(station_name: str, tn: str, slot: str) -> str:
            d = key_to_time.get((station_name, tn)) or key_to_time_norm.get((normalize(str(station_name)), tn)) or {}
            return d.get(slot, "")

        xlsx_dual_stations: set[tuple] = set()
        for (st_name, tn), times_d in key_to_time.items():
            if "o" in times_d:
                for s_name, s_km in stations_sorted:
                    if s_name == st_name or normalize(str(s_name)) == normalize(str(st_name)):
                        xlsx_dual_stations.add((s_name, s_km))
                        break

        row_offset = 0
        for i, (station_name, km_val) in enumerate(stations_sorted):
            if (station_name, km_val) in xlsx_dual_stations:
                r_p = start_row + i + row_offset
                c_km_p = ws.cell(row=r_p, column=4, value=float(km_val))
                try:
                    c_km_p.number_format = "0.000"
                except Exception:
                    pass
                ws.cell(row=r_p, column=5, value=str(station_name))
                ws.cell(row=r_p, column=6, value="p")
                for j, tn in enumerate(train_nums):
                    t_str = _get_time(station_name, tn, "p")
                    if t_str:
                        ws.cell(row=r_p, column=7 + j, value=t_str)
                row_offset += 1
                r_o = start_row + i + row_offset
                c_km_o = ws.cell(row=r_o, column=4, value=float(km_val))
                try:
                    c_km_o.number_format = "0.000"
                except Exception:
                    pass
                ws.cell(row=r_o, column=5, value=str(station_name))
                ws.cell(row=r_o, column=6, value="o")
                for j, tn in enumerate(train_nums):
                    t_str = _get_time(station_name, tn, "o")
                    if t_str:
                        ws.cell(row=r_o, column=7 + j, value=t_str)
            else:
                r = start_row + i + row_offset
                c_km = ws.cell(row=r, column=4, value=float(km_val))
                try:
                    c_km.number_format = "0.000"
                except Exception:
                    pass
                ws.cell(row=r, column=5, value=str(station_name))
                for j, tn in enumerate(train_nums):
                    t_str = _get_time(station_name, tn, "p")
                    if t_str:
                        ws.cell(row=r, column=7 + j, value=t_str)

        ws.cell(row=start_row + len(stations_sorted) + row_offset, column=5, value="do stacji")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_project_json(session: Any) -> bytes:
    project = {
        "_format": "train-timetable-plotter-project",
        "_version": 1,
        "uploaded_name": session.get("uploaded_name", ""),
        "selected_sheet": session.get("selected_sheet", ""),
        "station_map": session.get("station_map", {}),
        "station_maps": session.get("station_maps", {}),
        "sheets_data": session.get("sheets_data", []),
        "train_colors": session.get("train_colors", {}),
    }
    return json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")


def build_circuits_excel_bytes(session: Any) -> bytes:
    """Build XLSX with vehicle circuits (obiegi pojazdow) grouped by color."""
    _HEX_TO_NAME = {
        "#e6194b": "czerwony",
        "#4363d8": "niebieski",
        "#3cb44b": "zielony",
        "#f58231": "pomaranczowy",
        "#911eb4": "fioletowy",
        "#ffe119": "zolty",
    }

    sheets_data = session.get("sheets_data", [])
    train_colors = session.get("train_colors", {})

    all_trains: dict[str, list] = {}
    for entry in sheets_data:
        for rec in entry.get("trains", []):
            tn = str(rec.get("train_number"))
            all_trains.setdefault(tn, []).append(rec)

    train_summaries: dict[str, dict] = {}
    for tn, records in all_trains.items():
        timed = []
        for r in records:
            td = r.get("time_decimal")
            if td is not None:
                try:
                    timed.append((float(td), r))
                except (ValueError, TypeError):
                    pass
        if len(timed) < 2:
            continue
        timed.sort(key=lambda x: x[0])
        dep_time = timed[0][0]
        dep_station = timed[0][1].get("station", "")
        arr_time = timed[-1][0]
        arr_station = timed[-1][1].get("station", "")
        km_values = []
        for _, r in timed:
            try:
                km_values.append(float(r.get("km", 0)))
            except (ValueError, TypeError):
                pass
        km = abs(max(km_values) - min(km_values)) if km_values else 0.0
        train_summaries[tn] = {
            "dep_station": dep_station,
            "dep_time": dep_time,
            "arr_station": arr_station,
            "arr_time": arr_time,
            "km": km,
        }

    color_groups: dict[str, list] = {}
    unassigned: list = []
    for tn, summary in train_summaries.items():
        color = train_colors.get(tn)
        if color and color != "#000000":
            color_groups.setdefault(color, []).append((tn, summary))
        else:
            unassigned.append((tn, summary))

    for color in color_groups:
        color_groups[color].sort(key=lambda x: x[1]["dep_time"])
    unassigned.sort(key=lambda x: x[1]["dep_time"])

    sorted_groups = sorted(
        color_groups.items(),
        key=lambda item: min(s["dep_time"] for _, s in item[1]),
    )

    groups = []
    for color_hex, trains in sorted_groups:
        name = _HEX_TO_NAME.get(color_hex, color_hex)
        label = f"Obieg '{name}'"
        groups.append((label, trains))
    if unassigned:
        groups.append(("Obieg 'pociagi nieprzypisane'", unassigned))

    wb_c = Workbook()
    ws = wb_c.active
    ws.title = "Obiegi_pojazdow"

    _wfill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    _calign = Alignment(horizontal="center", vertical="center")
    _font = Font(name="Calibri", size=11)
    _thick = Side(style="thick", color="000000")
    _thin = Side(style="thin", color="AAAAAA")
    _no = Side()

    def _apply(cell: Any, left: Any = None, right: Any = None, top: Any = None, bottom: Any = None) -> None:
        cell.fill = _wfill
        cell.alignment = _calign
        cell.font = _font
        cell.border = Border(left=left or _no, right=right or _no,
                             top=top or _no, bottom=bottom or _no)

    ws.column_dimensions["E"].width = 19.57
    ws.column_dimensions["F"].width = 22.86
    ws.column_dimensions["I"].width = 19.43

    ws.row_dimensions[2].height = 16.5
    headers = ["Obieg", "Nr poc.", "Odj. RT", "Rel. od", "Rel. do",
                "Prz. RT", "Obsluga", "Uwagi", "km"]
    for i, h in enumerate(headers):
        col = 2 + i
        cell = ws.cell(row=2, column=col, value=h)
        if h in ("Odj. RT", "Prz. RT"):
            cell.number_format = "h:mm"
        _apply(cell,
               left=_thick if col == 2 else _thin,
               right=_thick if col == 10 else _thin,
               top=_thick, bottom=_thick)

    def _write_gap_rows(r_empty: int, r_sep: int) -> None:
        ws.row_dimensions[r_empty].height = 15.75
        ws.row_dimensions[r_sep].height = 19.5
        for col in range(2, 11):
            ce = ws.cell(row=r_empty, column=col)
            _apply(ce,
                   left=_thick if col == 2 else None,
                   right=_thick if col == 10 else None,
                   top=_thick)
            cs = ws.cell(row=r_sep, column=col)
            _apply(cs,
                   left=_thick if col == 2 else None,
                   right=_thick if col == 10 else None,
                   bottom=_thick)
            if col in (4, 7):
                cs.number_format = "h:mm"
        ws.cell(row=r_sep, column=3, value="")

    row_num = 2
    for label, trains in groups:
        row_num += 1
        r_empty = row_num
        row_num += 1
        r_sep = row_num
        _write_gap_rows(r_empty, r_sep)

        for t_idx, (tn, summary) in enumerate(trains):
            row_num += 1
            is_first = (t_idx == 0)
            ws.row_dimensions[row_num].height = 15.75

            b_cell = ws.cell(row=row_num, column=2,
                             value=label if is_first else None)
            _apply(b_cell, left=_thick,
                   top=_thick if is_first else None)

            ws.cell(row=row_num, column=3,
                    value=int(tn) if tn.isdigit() else tn)
            ws.cell(row=row_num, column=4,
                    value=_decimal_to_time(summary["dep_time"])).number_format = "h:mm"
            ws.cell(row=row_num, column=5, value=summary["dep_station"])
            ws.cell(row=row_num, column=6, value=summary["arr_station"])
            ws.cell(row=row_num, column=7,
                    value=_decimal_to_time(summary["arr_time"])).number_format = "h:mm"
            ws.cell(row=row_num, column=10, value=summary["km"])

            top_s = _thick if is_first else _thin
            for col in range(3, 11):
                _apply(ws.cell(row=row_num, column=col),
                       left=_thin,
                       right=_thick if col == 10 else None,
                       top=top_s, bottom=_thin)

    if groups:
        row_num += 1
        r_empty = row_num
        row_num += 1
        r_sep = row_num
        _write_gap_rows(r_empty, r_sep)

    buf = BytesIO()
    wb_c.save(buf)
    buf.seek(0)
    return buf.getvalue()
