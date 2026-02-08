import streamlit as st
import datetime as dt
import hashlib

from excel_loader import read_and_store_in_session
from table_editor import save_cell_time, clear_cell_time, propagate_time_shift
from utils import parse_time, format_time_hhmm
from train_grid_component.backend.train_grid_component import train_grid
from train_plot_component.backend.train_plot_component import train_plot

# Compat: st.dialog (Streamlit 1.37+) or st.experimental_dialog (1.34+)
_dialog_decorator = getattr(st, 'dialog', None) or getattr(st, 'experimental_dialog', None)


def _decimal_to_time(d: float) -> dt.time:
    """Convert decimal hours to dt.time, handling m==60 edge case."""
    h = int(d) % 24
    m = int(round((d % 1) * 60))
    if m == 60:
        h = (h + 1) % 24
        m = 0
    return dt.time(h, m)


st.set_page_config(layout="wide", page_title="wykresy z tabeli - KD")

# --- Tło i stylizacja strony ---
st.markdown("""
<style>
    /* Ciepłe kremowe tło z delikatnym wzorem lnianych linii */
    .stApp {
        background-color: #faf7f1;
        background-image:
            repeating-linear-gradient(
                45deg,
                rgba(195, 176, 145, 0.08),
                rgba(195, 176, 145, 0.08) 1px,
                transparent 1px,
                transparent 14px
            ),
            repeating-linear-gradient(
                -45deg,
                rgba(195, 176, 145, 0.08),
                rgba(195, 176, 145, 0.08) 1px,
                transparent 1px,
                transparent 14px
            );
    }

    /* Wymuszenie ciemnego tekstu niezależnie od trybu przeglądarki */
    .stApp, .stApp * {
        color: #1a1a1a;
    }

    /* Nagłówek */
    .stApp h1 {
        color: #4a3728;
    }

    /* Sidebar dopasowany kolorystycznie */
    [data-testid="stSidebar"] {
        background-color: #f3ece0;
    }
    [data-testid="stSidebar"], [data-testid="stSidebar"] * {
        color: #1a1a1a;
    }

    /* Kontenery z lekkim podkładem */
    [data-testid="stVerticalBlock"] > div[data-testid="stHorizontalBlock"],
    .stContainer {
        border-radius: 6px;
    }

    /* Upload i przycisk pobierania — jasne tło niezależnie od trybu przeglądarki */
    [data-testid="stFileUploader"],
    [data-testid="stFileUploader"] * {
        background-color: transparent;
    }
    [data-testid="stFileUploader"] section {
        background-color: #e8f0fe;
        border-color: #90b4e0;
    }
    [data-testid="stDownloadButton"] button {
        background-color: #e8f0fe;
        border-color: #90b4e0;
        color: #1a1a1a;
    }
    [data-testid="stDownloadButton"] button:hover {
        background-color: #d0e2fc;
        border-color: #6a9bd8;
    }
</style>
""", unsafe_allow_html=True)

st.title("Rozkład Jazdy - wykresy z tabeli")

uploaded_file = st.file_uploader("Prześlij plik Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        # zapamiętaj nazwę wczytanego pliku do nazwania eksportu
        try:
            st.session_state["uploaded_name"] = str(uploaded_file.name)
        except Exception:
            pass
        file_bytes = uploaded_file.getvalue()
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        if st.session_state.get("uploaded_hash") != file_hash:
            read_and_store_in_session(file_bytes, st.session_state)
            st.session_state["uploaded_hash"] = file_hash
            st.success("Dane wczytane i zapisane do st.session_state")
    except Exception as e:
        st.error(f"Nie udało się wczytać pliku: {e}")

"""Tabela km – stacja – pociągi"""
station_map = st.session_state.get("station_map", {})
sheets_data = st.session_state.get("sheets_data", [])

if station_map and sheets_data:
    # Selektor arkusza w obramowaniu z aktywnym stylem (radio poziome)
    sheet_names = [entry.get("sheet") for entry in sheets_data]
    if "selected_sheet" not in st.session_state:
        st.session_state.selected_sheet = sheet_names[0] if sheet_names else None

    with st.container(border=True):
        st.markdown("**Arkusz:**")
        selected_sheet = st.radio(
            label="Arkusz",
            label_visibility="collapsed",
            options=sheet_names,
            horizontal=True,
            index=(sheet_names.index(st.session_state.selected_sheet)
                   if st.session_state.selected_sheet in sheet_names else 0),
            key="sheet_radio",
        )
    st.session_state.selected_sheet = selected_sheet

    # Filtr danych do wybranego arkusza
    active = next((entry for entry in sheets_data if entry.get("sheet") == selected_sheet), {"trains": []})
    trains_active = active.get("trains", [])

    # Przyciski akcji
    with st.container(border=True):
        col_dl, _ = st.columns([1,5])

        def _hhmm_from_any(val: any) -> str:
            try:
                if isinstance(val, (int, float)):
                    d = float(val)
                else:
                    d = parse_time(val)
                if d is None:
                    return ""
                return format_time_hhmm(d)
            except Exception:
                return ""

        def build_excel_bytes() -> bytes:
            from io import BytesIO
            from openpyxl import Workbook
            from utils import normalize

            wb = Workbook()
            # usuń domyślny arkusz
            try:
                wb.remove(wb.active)
            except Exception:
                pass

            station_maps_all = st.session_state.get("station_maps", {})

            for entry in sheets_data:
                sheet_name = str(entry.get("sheet"))
                ws = wb.create_sheet(title=sheet_name[:31] or "Arkusz")

                # Nagłówki stałe
                ws["E3"] = "numer pociągu"
                ws["D11"] = "km"
                ws["E11"] = "ze stacji"
                ws["F11"] = "p/o"

                # Lista stacji/km z mapy wybranego arkusza (dla każdego arkusza jego własna mapa)
                station_map_sheet = station_maps_all.get(sheet_name, {})
                stations_sorted = sorted(station_map_sheet.items(), key=lambda kv: kv[1])

                start_row = 12

                # Kolumny pociągów – unikalne numery
                trains_list = entry.get("trains", [])
                train_nums = list(dict.fromkeys(str(t.get("train_number")) for t in trains_list))
                # wiersz 3 od kolumny G w prawo (kolumna F = p/o)
                for j, tn in enumerate(train_nums):
                    ws.cell(row=3, column=7 + j, value=tn)

                # Zbuduj mapę czasów: (station, tn) -> {"p": time, "o": time}
                # Preferuj time_decimal do formatowania
                key_to_time = {}
                key_to_time_norm = {}
                for rec in trains_list:
                    st_name = rec.get("station")
                    tn = str(rec.get("train_number"))
                    t_fmt = _hhmm_from_any(rec.get("time_decimal") if rec.get("time_decimal") is not None else rec.get("time"))
                    rec_stop_type = rec.get("stop_type", "p")
                    key_to_time.setdefault((st_name, tn), {})[rec_stop_type] = t_fmt
                    key_to_time_norm.setdefault((normalize(str(st_name)), tn), {})[rec_stop_type] = t_fmt

                def _get_time(station_name, tn, slot):
                    d = key_to_time.get((station_name, tn)) or key_to_time_norm.get((normalize(str(station_name)), tn)) or {}
                    return d.get(slot, "")

                # Wykryj stacje z podwójnym wpisem (przyjazd + odjazd)
                xlsx_dual_stations = set()
                for (st_name, tn), times_d in key_to_time.items():
                    if "o" in times_d:
                        for s_name, s_km in stations_sorted:
                            if s_name == st_name or normalize(str(s_name)) == normalize(str(st_name)):
                                xlsx_dual_stations.add((s_name, s_km))
                                break

                # Wypełnij czasy wg stacji i numerów pociągów (z obsługą podwójnych wierszy)
                row_offset = 0
                for i, (station_name, km_val) in enumerate(stations_sorted):
                    if (station_name, km_val) in xlsx_dual_stations:
                        # Wiersz przyjazdu (p)
                        r_p = start_row + i + row_offset
                        c_km_p = ws.cell(row=r_p, column=4, value=float(km_val))
                        try:
                            c_km_p.number_format = "0.000"
                        except Exception:
                            pass
                        ws.cell(row=r_p, column=5, value=str(station_name))
                        ws.cell(row=r_p, column=6, value="p")
                        for j, tn in enumerate(train_nums):
                            t_str = _get_time(station_name, tn, "p")
                            if t_str:
                                ws.cell(row=r_p, column=7 + j, value=t_str)
                        # Wiersz odjazdu (o)
                        row_offset += 1
                        r_o = start_row + i + row_offset
                        c_km_o = ws.cell(row=r_o, column=4, value=float(km_val))
                        try:
                            c_km_o.number_format = "0.000"
                        except Exception:
                            pass
                        ws.cell(row=r_o, column=5, value=str(station_name))
                        ws.cell(row=r_o, column=6, value="o")
                        for j, tn in enumerate(train_nums):
                            t_str = _get_time(station_name, tn, "o")
                            if t_str:
                                ws.cell(row=r_o, column=7 + j, value=t_str)
                    else:
                        r = start_row + i + row_offset
                        c_km = ws.cell(row=r, column=4, value=float(km_val))
                        try:
                            c_km.number_format = "0.000"
                        except Exception:
                            pass
                        ws.cell(row=r, column=5, value=str(station_name))
                        for j, tn in enumerate(train_nums):
                            t_str = _get_time(station_name, tn, "p")
                            if t_str:
                                ws.cell(row=r, column=7 + j, value=t_str)

                # "do stacji" w wierszu po ostatniej stacji (uwzględnia dodatkowe wiersze)
                ws.cell(row=start_row + len(stations_sorted) + row_offset, column=5, value="do stacji")

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.getvalue()

        with col_dl:
            xbytes = build_excel_bytes()
            export_name = st.session_state.get("uploaded_name") or "rozkład.xlsx"
            st.download_button(
                label="Pobierz rozkład do xlsx",
                data=xbytes,
                file_name=export_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # --- Narzędzie koloru pociągów ---
    if "train_colors" not in st.session_state:
        st.session_state["train_colors"] = {}
    if "active_color" not in st.session_state:
        st.session_state["active_color"] = None

    _COLOR_PALETTE = [
        ("Czerwony", "#e6194b"),
        ("Niebieski", "#4363d8"),
        ("Zielony", "#3cb44b"),
        ("Pomarańczowy", "#f58231"),
        ("Fioletowy", "#911eb4"),
        ("Żółty", "#ffe119"),
        ("Czarny", "#000000"),
    ]

    with st.container(border=True):
        st.markdown("**Narzędzie koloru**")
        _cc = st.columns(len(_COLOR_PALETTE) + 2)
        for idx, (name, hexc) in enumerate(_COLOR_PALETTE):
            with _cc[idx]:
                _swatch = f'<div style="width:100%;height:8px;background:{hexc};border-radius:3px;margin-bottom:4px"></div>'
                st.markdown(_swatch, unsafe_allow_html=True)
                _btn_type = "primary" if st.session_state["active_color"] == hexc else "secondary"
                if st.button(name, key=f"color_btn_{hexc}", type=_btn_type):
                    st.session_state["active_color"] = hexc
                    st.rerun()
        # "Brak koloru" — deaktywacja narzędzia
        with _cc[len(_COLOR_PALETTE)]:
            _btn_type_none = "primary" if st.session_state["active_color"] is None else "secondary"
            if st.button("Brak koloru", key="color_btn_none", type=_btn_type_none):
                st.session_state["active_color"] = None
                st.rerun()
        # "Wyczyść kolory" — reset wszystkich kolorów
        with _cc[len(_COLOR_PALETTE) + 1]:
            if st.button("Wyczyść kolory", key="color_btn_clear"):
                st.session_state["train_colors"] = {}
                st.session_state["active_color"] = None
                # Incrementuj nonce żeby odświeżyć komponenty
                for _nk in list(st.session_state.keys()):
                    if _nk.startswith("plot_nonce_") or _nk.startswith("grid_nonce_"):
                        st.session_state[_nk] += 1
                st.rerun()

    _active_color = st.session_state["active_color"]
    _train_colors = st.session_state["train_colors"]

    # Zbuduj kolumny pociągów (tylko z wybranego arkusza)
    train_numbers = [str(t["train_number"]) for t in trains_active]
    unique_trains = list(dict.fromkeys(train_numbers))

    # Wiersze bazujemy na mapie stacji z wybranego arkusza, posortowane po km rosnąco
    station_maps = st.session_state.get("station_maps", {})
    active_station_map = station_maps.get(selected_sheet, station_map)
    station_items = sorted(active_station_map.items(), key=lambda kv: kv[1])

    # Zbuduj strukturę: {(station, km): {train_number: {"p": val, "o": val}}} tylko dla wybranego arkusza
    cell_map = {}
    for rec in trains_active:
        key = (rec["station"], float(rec["km"]))
        bucket = cell_map.setdefault(key, {})
        # Preferuj time_decimal, aby ujednolicić wyświetlanie bez sufiksu (+d)
        try:
            tdec = float(rec.get("time_decimal")) if rec.get("time_decimal") is not None else None
        except Exception:
            tdec = None
        if tdec is None:
            parsed = parse_time(rec.get("time"))
            tdec = float(parsed) if parsed is not None else None
        if tdec is not None:
            display_val = format_time_hhmm(tdec)
        else:
            display_val = str(rec.get("time") or "")
        stop_type = rec.get("stop_type", "p")  # non-dual defaults to "p"
        bucket.setdefault(str(rec["train_number"]), {})[stop_type] = display_val

    # Wykryj stacje z podwójnym wpisem (przyjazd + odjazd)
    dual_stations = set()
    for key, times_dict in cell_map.items():
        if any("o" in v for v in times_dict.values()):
            dual_stations.add(key)

    # Render: tabela
    import pandas as pd
    table_rows = []
    for station, km in ((name, km) for name, km in station_items):
        times = cell_map.get((station, float(km)), {})
        if (station, float(km)) in dual_stations:
            # Wiersz przyjazdu (p)
            row_p = {"km": f"{km:.3f}", "stacja": f"{station} (p)", "_station_raw": station, "_stop_type": "p"}
            for tn in unique_trains:
                row_p[tn] = times.get(tn, {}).get("p", "")
            table_rows.append(row_p)
            # Wiersz odjazdu (o)
            row_o = {"km": f"{km:.3f}", "stacja": f"{station} (o)", "_station_raw": station, "_stop_type": "o"}
            for tn in unique_trains:
                row_o[tn] = times.get(tn, {}).get("o", "")
            table_rows.append(row_o)
        else:
            row = {"km": f"{km:.3f}", "stacja": station, "_station_raw": station, "_stop_type": None}
            for tn in unique_trains:
                row[tn] = times.get(tn, {}).get("p", "")
            table_rows.append(row)

    df_view = pd.DataFrame(table_rows)

    # Wykres tras (nad tabelą)
    st.subheader("Wykres tras pociągów")
    plot_height = st.slider("Wysokość wykresu", min_value=600, max_value=4000, value=600, step=100, key="plot_height")
    # Oś Y: km (zaznaczamy stacje z mapy aktywnego arkusza jako markery)
    y_stations = [{"name": name, "km": float(km)} for name, km in station_items]

    # Budowa serii: dla każdego pociągu z KAŻDEGO arkusza punkty [czas_ms, indeks_stacji]
    def hours_to_ms(hours_float: float) -> int:
        return int(float(hours_float) * 3600000.0)

    # Zbuduj dla każdego arkusza mapę: station -> {train_number: time_decimal}
    # time_decimal is already midnight-corrected at load time, so no inline correction needed.
    sheet_to_station_train = {}
    sheet_to_trains = {}
    for entry in sheets_data:
        sheet = entry.get("sheet")
        station_to_train = {}
        trains = entry.get("trains", [])
        for rec in trains:
            station_name = rec["station"]
            tn = str(rec["train_number"])
            station_bucket = station_to_train.setdefault(station_name, {})
            station_bucket.setdefault(tn, []).append(rec.get("time_decimal"))
        sheet_to_station_train[sheet] = station_to_train
        sheet_to_trains[sheet] = sorted({ str(t["train_number"]) for t in trains })

    # Dla osi Y wykorzystujemy km z wybranego arkusza; rysujemy WSZYSTKIE pociągi z KAŻDEGO arkusza
    series = []
    global_min_ms = None
    global_max_ms = None
    for sheet, station_to_train in sheet_to_station_train.items():
        for tn in sheet_to_trains.get(sheet, []):
            pts = []
            for station_name, km_selected in station_items:
                times_list = station_to_train.get(station_name, {}).get(tn)
                if not times_list:
                    continue
                for t_dec in times_list:
                    if t_dec is None:
                        continue
                    ms = hours_to_ms(t_dec)
                    pts.append({
                        "value": [ms, float(km_selected)],
                        "station": station_name,
                        "train": tn,
                        "sheet": sheet,
                    })
                    global_min_ms = ms if global_min_ms is None else min(global_min_ms, ms)
                    global_max_ms = ms if global_max_ms is None else max(global_max_ms, ms)
            if pts:
                # Sortuj punkty po czasie, aby linia nie cofała się przy stacjach dualnych
                pts.sort(key=lambda p: p["value"][0])
                # Nazwa serii zawiera arkusz, by rozróżnić te same numery pociągów w różnych arkuszach
                series.append({"name": f"{tn} ({sheet})", "points": pts})

    # Dynamiczny zakres X z marginesem: lewy -2h dla etykiet, prawy +30 min
    pad_left = 2 * 60 * 60 * 1000
    pad_right = 30 * 60 * 1000
    x_min = max(0, (global_min_ms or 0) - pad_left)
    x_max = (global_max_ms or (24 * 3600 * 1000)) + pad_right

    # Nonce do resetu komponentu wykresu i siatki po akcji w modalu
    plot_nonce_key = f"plot_nonce_{selected_sheet}"
    if plot_nonce_key not in st.session_state:
        st.session_state[plot_nonce_key] = 0
    grid_nonce_key = f"grid_nonce_{selected_sheet}"
    if grid_nonce_key not in st.session_state:
        st.session_state[grid_nonce_key] = 0

    evt_plot = train_plot(
        y_stations=y_stations,
        series=series,
        x_min_ms=x_min,
        x_max_ms=x_max,
        key=f"plot_{selected_sheet}_{st.session_state[plot_nonce_key]}",
        height=int(plot_height or 600),
        train_colors=_train_colors,
        color_mode=(_active_color is not None),
    )


    # Obsługa single-click z wykresu (tryb koloru)
    if evt_plot and isinstance(evt_plot, dict) and evt_plot.get("type") == "pointClick" and _active_color is not None:
        _click_train = str(evt_plot.get("train") or "")
        if _click_train:
            if _active_color == "#000000":
                _train_colors.pop(_click_train, None)
            else:
                _train_colors[_click_train] = _active_color
            st.session_state["train_colors"] = _train_colors
            st.session_state[plot_nonce_key] += 1
            st.session_state[grid_nonce_key] += 1
            st.rerun()

    # Obsługa dblclick z wykresu (edycja czasu w dowolnym arkuszu)
    if evt_plot and isinstance(evt_plot, dict) and evt_plot.get("type") == "pointDoubleClick" and not _active_color:
        try:
            col_id = str(evt_plot.get("train") or "")
            station_clicked = str(evt_plot.get("station") or "")
            km_clicked = float(evt_plot.get("km") or 0.0)
            # arkusz może być None — wtedy użyj aktualnego
            sheet_clicked = str(evt_plot.get("sheet") or selected_sheet)
        except Exception:
            col_id, station_clicked, km_clicked, sheet_clicked = "", "", 0.0, selected_sheet

        # km dla propagacji i zapisu bierzemy z mapy stacji właściwego arkusza
        try:
            km_sheet_clicked = float(st.session_state.get("station_maps", {}).get(sheet_clicked, {}).get(station_clicked, km_clicked))
        except Exception:
            km_sheet_clicked = km_clicked

        # Domyślna godzina z eventu (ms od północy, z obsługą doby >24h)
        try:
            ms_val = float(evt_plot.get("ms"))
            parsed = ms_val / 3600000.0
        except Exception:
            parsed = None
        default_time = _decimal_to_time(parsed) if parsed is not None else dt.time(0, 0)

        _day_offset_plot = int(parsed // 24) if parsed is not None else 0

        # Determine stop_type by matching clicked time to existing records
        _stop_type_plot = None
        if parsed is not None:
            _plot_sheet_data = next((s for s in sheets_data if s.get("sheet") == sheet_clicked), None)
            if _plot_sheet_data:
                _best_delta = None
                for _rec in _plot_sheet_data.get("trains", []):
                    if (str(_rec.get("train_number")) == col_id
                            and _rec.get("station") == station_clicked
                            and abs(float(_rec.get("km", 0)) - km_sheet_clicked) < 0.01):
                        _rd = _rec.get("time_decimal")
                        if _rd is not None:
                            _delta = abs(float(_rd) - parsed)
                            if _best_delta is None or _delta < _best_delta:
                                _best_delta = _delta
                                _stop_type_plot = _rec.get("stop_type")

        try:
            @_dialog_decorator("Edycja czasu (z wykresu)")
            def time_dialog_plot():
                st.write(f"Arkusz: {sheet_clicked}  •  Stacja: {station_clicked}  •  km: {km_sheet_clicked:.3f}  •  Pociąg: {col_id}")
                t = st.time_input("Godzina", value=default_time, step=dt.timedelta(minutes=1), key=f"dlg_time_plot_{sheet_clicked}")
                allow_propagate = parsed is not None
                prop = st.checkbox("Uwzględnij zmianę na dalszej części trasy", value=True, disabled=not allow_propagate, key=f"dlg_prop_plot_{sheet_clicked}")
                c1, c2, c3 = st.columns([1,1,1])
                with c1:
                    if st.button("Zapisz", type="primary", key=f"dlg_save_plot_{sheet_clicked}"):
                        if prop and parsed is not None:
                            new_dec = float(t.hour) + float(t.minute)/60.0 + float(getattr(t, 'second', 0))/3600.0
                            parsed_norm = float(parsed) % 24
                            delta_hours = new_dec - parsed_norm
                        else:
                            delta_hours = 0.0
                        save_cell_time(sheet_clicked, station_clicked, float(km_sheet_clicked), col_id, t, st.session_state, day_offset=_day_offset_plot, stop_type=_stop_type_plot)
                        if prop and delta_hours != 0.0:
                            propagate_time_shift(sheet_clicked, col_id, float(km_sheet_clicked), float(delta_hours), st.session_state)
                        st.session_state[plot_nonce_key] += 1
                        st.session_state[grid_nonce_key] += 1
                        st.rerun()
                with c2:
                    if st.button("Usuń postój", key=f"dlg_clear_plot_{sheet_clicked}"):
                        clear_cell_time(sheet_clicked, station_clicked, float(km_sheet_clicked), col_id, st.session_state, stop_type=_stop_type_plot)
                        st.session_state[plot_nonce_key] += 1
                        st.session_state[grid_nonce_key] += 1
                        st.rerun()
                with c3:
                    if st.button("Anuluj", key=f"dlg_cancel_plot_{sheet_clicked}"):
                        st.session_state[plot_nonce_key] += 1
                        st.session_state[grid_nonce_key] += 1
                        st.rerun()

            time_dialog_plot()
        except Exception as e:
            # Fallback: panel zamiast modala
            st.warning(f"Nie udało się otworzyć okna dialogowego wykresu: {e}")
            with st.container(border=True):
                st.write(f"Arkusz: {sheet_clicked}  •  Stacja: {station_clicked}  •  km: {km_sheet_clicked:.3f}  •  Pociąg: {col_id}")
                t = st.time_input("Godzina", value=default_time, step=dt.timedelta(minutes=1), key=f"fallback_time_plot_{sheet_clicked}")
                allow_propagate_fb = parsed is not None
                prop_fb = st.checkbox("Uwzględnij zmianę na dalszej części trasy", value=True, disabled=not allow_propagate_fb, key=f"fallback_prop_plot_{sheet_clicked}")
                c1, c2, c3 = st.columns([1,1,1])
                with c1:
                    if st.button("Zapisz", type="primary", key=f"fallback_save_plot_{sheet_clicked}"):
                        if prop_fb and parsed is not None:
                            new_dec = float(t.hour) + float(t.minute)/60.0 + float(getattr(t, 'second', 0))/3600.0
                            parsed_norm = float(parsed) % 24
                            delta_hours = new_dec - parsed_norm
                        else:
                            delta_hours = 0.0
                        save_cell_time(sheet_clicked, station_clicked, float(km_sheet_clicked), col_id, t, st.session_state, day_offset=_day_offset_plot, stop_type=_stop_type_plot)
                        if prop_fb and delta_hours != 0.0:
                            propagate_time_shift(sheet_clicked, col_id, float(km_sheet_clicked), float(delta_hours), st.session_state)
                        st.session_state[plot_nonce_key] += 1
                        st.session_state[grid_nonce_key] += 1
                        st.rerun()
                with c2:
                    if st.button("Usuń postój", key=f"fallback_clear_plot_{sheet_clicked}"):
                        clear_cell_time(sheet_clicked, station_clicked, float(km_sheet_clicked), col_id, st.session_state, stop_type=_stop_type_plot)
                        st.session_state[plot_nonce_key] += 1
                        st.session_state[grid_nonce_key] += 1
                        st.rerun()
                with c3:
                    if st.button("Anuluj", key=f"fallback_cancel_plot_{sheet_clicked}"):
                        st.session_state[plot_nonce_key] += 1
                        st.session_state[grid_nonce_key] += 1
                        st.rerun()

    st.subheader("Tabela: km – stacja – pociągi")
    st.caption(f"Arkusz: {selected_sheet}")

    # Zbuduj columnDefs i wywołaj custom component
    import pandas as pd
    train_cols = [c for c in df_view.columns if c not in ("km", "stacja", "_station_raw", "_stop_type")]
    column_defs = (
        [
            {"field": "km", "headerName": "km", "editable": False, "width": 60},
            {"field": "stacja", "headerName": "stacja", "editable": False, "width": 60},
        ]
        + [{"field": c, "headerName": c, "editable": True, "width": 120} for c in train_cols]
    )

    row_data = df_view.to_dict(orient="records")
    grid_height = min(600, 100 + 26 * (len(df_view) + 1))

    evt = train_grid(
        row_data=row_data,
        column_defs=column_defs,
        key=f"grid_{selected_sheet}_{st.session_state[grid_nonce_key]}",
        height=grid_height,
        theme="ag-theme-alpine",
        train_colors=_train_colors,
        color_mode=(_active_color is not None),
    )

    # Obsługa single-click z tabeli (tryb koloru)
    if evt and isinstance(evt, dict) and evt.get("type") == "cellClick" and _active_color is not None:
        _click_field = str(evt.get("field") or "")
        if _click_field and _click_field not in ("km", "stacja", "_station_raw", "_stop_type"):
            if _active_color == "#000000":
                _train_colors.pop(_click_field, None)
            else:
                _train_colors[_click_field] = _active_color
            st.session_state["train_colors"] = _train_colors
            st.session_state[plot_nonce_key] += 1
            st.session_state[grid_nonce_key] += 1
            st.rerun()

    # Obsługa eventów z komponentu
    if evt and isinstance(evt, dict) and evt.get("type") == "cellDoubleClick" and not _active_color:
        col_id = str(evt.get("field") or "")
        if col_id and col_id not in ("km", "stacja", "_station_raw", "_stop_type"):
            row = evt.get("row") or {}
            # Użyj ukrytych pól do identyfikacji stacji i typu postoju
            station_clicked = str(row.get("_station_raw") or row.get("stacja") or "")
            _stop_type_clicked = row.get("_stop_type")  # "p", "o", or None
            try:
                km_clicked = float(str(row.get("km") or "0").replace(",", "."))
            except Exception:
                km_clicked = 0.0

            # Ustal domyślną godzinę z danych (cell_map przechowuje dict z kluczami "p"/"o")
            _lookup_slot = _stop_type_clicked if _stop_type_clicked else "p"
            current_time_str = cell_map.get((station_clicked, float(km_clicked)), {}).get(col_id, {}).get(_lookup_slot, "")
            parsed = parse_time(current_time_str) if current_time_str else None
            default_time = _decimal_to_time(parsed) if parsed is not None else dt.time(0, 0)

            _day_offset_grid = int(parsed // 24) if parsed is not None else 0

            # Etykieta stacji do wyświetlenia w dialogu
            _station_label = str(row.get("stacja") or station_clicked)

            # Modal dialog edycji czasu (wymaga Streamlit 1.34+)
            try:
                @_dialog_decorator("Edycja czasu")
                def time_dialog():
                    st.write(f"Stacja: {_station_label}  •  km: {km_clicked:.3f}  •  Pociąg: {col_id}")
                    t = st.time_input("Godzina", value=default_time, step=dt.timedelta(minutes=1), key=f"dlg_time_{selected_sheet}")
                    # Checkbox propagacji tylko jeśli istniała poprzednia wartość czasu (mamy parsed)
                    allow_propagate = parsed is not None
                    prop = st.checkbox("Uwzględnij zmianę na dalszej części trasy", value=True, disabled=not allow_propagate, key=f"dlg_prop_{selected_sheet}")
                    c1, c2, c3 = st.columns([1,1,1])
                    with c1:
                        if st.button("Zapisz", type="primary", key=f"dlg_save_{selected_sheet}"):
                            # Oblicz delta względem poprzedniej wartości jeśli propagacja aktywna
                            if prop and parsed is not None:
                                new_dec = float(t.hour) + float(t.minute)/60.0 + float(getattr(t, 'second', 0))/3600.0
                                parsed_norm = float(parsed) % 24
                                delta_hours = new_dec - parsed_norm
                            else:
                                delta_hours = 0.0

                            save_cell_time(selected_sheet, station_clicked, float(km_clicked), col_id, t, st.session_state, day_offset=_day_offset_grid, stop_type=_stop_type_clicked)
                            if prop and delta_hours != 0.0:
                                propagate_time_shift(selected_sheet, col_id, float(km_clicked), float(delta_hours), st.session_state)
                            st.session_state[grid_nonce_key] += 1
                            st.session_state[plot_nonce_key] += 1
                            st.rerun()
                    with c2:
                        if st.button("Usuń postój", key=f"dlg_clear_{selected_sheet}"):
                            clear_cell_time(selected_sheet, station_clicked, float(km_clicked), col_id, st.session_state, stop_type=_stop_type_clicked)
                            st.session_state[grid_nonce_key] += 1
                            st.session_state[plot_nonce_key] += 1
                            st.rerun()
                    with c3:
                        if st.button("Anuluj", key=f"dlg_cancel_{selected_sheet}"):
                            st.session_state[grid_nonce_key] += 1
                            st.session_state[plot_nonce_key] += 1
                            st.rerun()

                time_dialog()
            except Exception as e:
                # Fallback: panel zamiast modala
                st.warning(f"Nie udało się otworzyć okna dialogowego: {e}")
                with st.container(border=True):
                    st.write(f"Stacja: {_station_label}  •  km: {km_clicked:.3f}  •  Pociąg: {col_id}")
                    t = st.time_input("Godzina", value=default_time, step=dt.timedelta(minutes=1), key=f"fallback_time_{selected_sheet}")
                    allow_propagate_fb = parsed is not None
                    prop_fb = st.checkbox("Uwzględnij zmianę na dalszej części trasy", value=True, disabled=not allow_propagate_fb, key=f"fallback_prop_{selected_sheet}")
                    c1, c2, c3 = st.columns([1,1,1])
                    with c1:
                        if st.button("Zapisz", type="primary", key=f"fallback_save_{selected_sheet}"):
                            if prop_fb and parsed is not None:
                                new_dec = float(t.hour) + float(t.minute)/60.0 + float(getattr(t, 'second', 0))/3600.0
                                parsed_norm = float(parsed) % 24
                                delta_hours = new_dec - parsed_norm
                            else:
                                delta_hours = 0.0
                            save_cell_time(selected_sheet, station_clicked, float(km_clicked), col_id, t, st.session_state, day_offset=_day_offset_grid, stop_type=_stop_type_clicked)
                            if prop_fb and delta_hours != 0.0:
                                propagate_time_shift(selected_sheet, col_id, float(km_clicked), float(delta_hours), st.session_state)
                            st.session_state[grid_nonce_key] += 1
                            st.session_state[plot_nonce_key] += 1
                            st.rerun()
                    with c2:
                        if st.button("Usuń postój", key=f"fallback_clear_{selected_sheet}"):
                            clear_cell_time(selected_sheet, station_clicked, float(km_clicked), col_id, st.session_state, stop_type=_stop_type_clicked)
                            st.session_state[grid_nonce_key] += 1
                            st.session_state[plot_nonce_key] += 1
                            st.rerun()
                    with c3:
                        if st.button("Anuluj", key=f"fallback_cancel_{selected_sheet}"):
                            st.session_state[grid_nonce_key] += 1
                            st.session_state[plot_nonce_key] += 1
                            st.rerun()
else:
    st.info("Brak danych do zbudowania tabeli. Wczytaj plik.")

st.markdown("---")

# Debugowy podgląd danych zapisanych w stanie aplikacji

st.title("🧪 Debug: Excel → session state")

st.subheader("Mapa stacji (z 1. arkusza)")
if station_map:
    st.json(station_map)
else:
    st.info("Brak danych. Wczytaj plik.")

st.subheader("Weryfikacja spójności listy stacji we wszystkich arkuszach")
station_check = st.session_state.get("station_check", {"ok": False, "mismatches": []})
st.write("Zgodność:", "TAK" if station_check.get("ok") else "NIE")
if station_check.get("mismatches"):
    st.json(station_check["mismatches"])

st.subheader("Dane pociągów per arkusz")
if sheets_data:
    for entry in sheets_data:
        st.markdown(f"**Arkusz:** {entry.get('sheet')}")
        trains = entry.get("trains", [])
        if not trains:
            st.write("(brak danych pociągów)")
            continue
        minimal_view = [
            {
                "train_number": t["train_number"],
                "station": t["station"],
                "km": t["km"],
                "time": t["time"],
            }
            for t in trains
        ]
        st.dataframe(minimal_view, use_container_width=True)
else:
    st.info("Brak danych pociągów. Wczytaj plik.")

# Footer – profesjonalny podpis
footer_year = dt.date.today().year
st.markdown(
    f"""
    <style>
    .app-footer {{ position: fixed; left: 0; right: 0; bottom: 0; padding: 6px 10px; text-align: center; color: #8a7560; background: rgba(243, 236, 224, 0.85); font-size: 12px; z-index: 10000; border-top: 1px solid rgba(195, 176, 145, 0.2); }}
    </style>
    <div class=\"app-footer\">© {footer_year} Kacper Szmajda</div>
    """,
    unsafe_allow_html=True,
)

